# -*- coding: utf-8 -*-
"""
====================================================================
 AGENTE DE TRATAMENTO DE DADOS — Base de Vendas Porsche
====================================================================

Este script funciona como um "agente" de limpeza de dados: recebe a
base bruta (com campos de texto livre, inconsistentes e mal formatados)
e devolve uma base tratada, padronizada e pronta para alimentar o
dashboard de vendas.

O agente é dividido em "sub-agentes" especializados — uma função por
tipo de campo — cada um documentando a lógica de decisão aplicada.
Sempre que o agente não consegue resolver um valor com confiança
(ex.: uma data impossível como 30/02/2024), ele marca o registro como
INVÁLIDO / PENDENTE DE REVISÃO em vez de "inventar" um valor.

Como a base de origem já traz colunas de referência (*Sanitized*)
geradas manualmentente para fins didáticos, o agente também roda uma
etapa de VALIDAÇÃO comparando sua própria limpeza com esse gabarito,
reportando a taxa de acerto por coluna.
"""

import re
import json
import unicodedata
from datetime import date, datetime

import openpyxl

# --------------------------------------------------------------------------
# Utilidades gerais
# --------------------------------------------------------------------------

def strip_accents(s):
    return "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")


# Pequeno parser de números por extenso em inglês (cobre o vocabulário
# necessário para preços e quilometragens escritos por extenso, como
# "eighty two thousand" ou "fifteen thousand miles").
_ONES = {
    "zero": 0, "one": 1, "two": 2, "three": 3, "four": 4, "five": 5, "six": 6,
    "seven": 7, "eight": 8, "nine": 9, "ten": 10, "eleven": 11, "twelve": 12,
    "thirteen": 13, "fourteen": 14, "fifteen": 15, "sixteen": 16,
    "seventeen": 17, "eighteen": 18, "nineteen": 19,
}
_TENS = {
    "twenty": 20, "thirty": 30, "forty": 40, "fifty": 50, "sixty": 60,
    "seventy": 70, "eighty": 80, "ninety": 90,
}
_SCALES = {"hundred": 100, "thousand": 1000, "million": 1_000_000}


def words_to_number(text):
    """Converte um número escrito por extenso em inglês para int.
    Retorna None se a string não for reconhecida como número por extenso."""
    words = re.findall(r"[a-z]+", text.lower())
    if not words or not all(w in _ONES or w in _TENS or w in _SCALES or w == "and" for w in words):
        return None
    total = 0
    current = 0
    for w in words:
        if w == "and":
            continue
        elif w in _ONES:
            current += _ONES[w]
        elif w in _TENS:
            current += _TENS[w]
        elif w == "hundred":
            current *= 100
        elif w in ("thousand", "million"):
            current = max(current, 1) * _SCALES[w]
            total += current
            current = 0
    return total + current


# --------------------------------------------------------------------------
# 1) Data da venda
# --------------------------------------------------------------------------

def clean_date(raw, min_year=2020, max_year=2027):
    """Interpreta datas em múltiplos formatos (MM/DD/AAAA, AAAA-MM-DD,
    separadores '-', '.', '/'), assumindo padrão norte-americano
    (mês/dia/ano) para datas ambíguas de 2 ou 4 dígitos.
    Retorna uma string ISO 'AAAA-MM-DD', ou 'INVALID' quando a data
    não existe no calendário (ex.: 30/02) ou não pode ser interpretada
    com segurança.
    """
    if raw is None:
        return "INVALID"
    if isinstance(raw, datetime):
        return raw.strftime("%Y-%m-%d")
    if isinstance(raw, date):
        return raw.strftime("%Y-%m-%d")

    s = str(raw).strip()
    if not s:
        return "INVALID"

    # Formatos textuais tipo "April 31st, 2024" ou "Dec 25th 2024"
    m = re.match(r"([A-Za-z]+)\.?\s+(\d{1,2})(?:st|nd|rd|th)?,?\s+(\d{4})", s)
    if m:
        month_name, day_s, year_s = m.groups()
        months = ["january", "february", "march", "april", "may", "june", "july",
                  "august", "september", "october", "november", "december"]
        months_abbr = [m[:3] for m in months]
        month_l = month_name.lower()
        if month_l in months:
            month = months.index(month_l) + 1
            return _build_date(int(year_s), month, int(day_s))
        if month_l in months_abbr:
            month = months_abbr.index(month_l) + 1
            return _build_date(int(year_s), month, int(day_s))
        return "INVALID"

    # Formato ISO: AAAA-MM-DD ou AAAA.MM.DD ou AAAA/MM/DD
    # (a posição do meio é sempre o MÊS; se exceder 12, a data é
    # considerada irrecuperável — não tentamos "adivinhar" invertendo
    # dia/mês, pois isso mascararia entradas realmente malformadas)
    m = re.match(r"^(\d{4})[-./](\d{1,2})[-./](\d{1,2})", s)
    if m:
        year_s, month_s, day_s = m.groups()
        return _build_date(int(year_s), int(month_s), int(day_s))

    # Formato MM/DD/AAAA, MM-DD-AAAA, MM.DD.AAAA (ano com 2 ou 4 dígitos)
    m = re.match(r"^(\d{1,2})[-./](\d{1,2})[-./](\d{2,4})$", s)
    if m:
        month_s, day_s, year_s = m.groups()
        month, day = int(month_s), int(day_s)
        year = int(year_s)
        if year < 100:
            year += 2000
        return _build_date(year, month, day)

    return "INVALID"


def _build_date(year, month, day):
    try:
        d = date(year, month, day)
    except ValueError:
        return "INVALID"
    if not (2015 <= d.year <= 2035):
        return "INVALID"
    return d.strftime("%Y-%m-%d")


# --------------------------------------------------------------------------
# 2) Ano do modelo
# --------------------------------------------------------------------------

# Mapeamento direto para as frases por extenso observadas na base
# (cobre o padrão "twenty twenty-four" = dois blocos de dois dígitos,
# e "two thousand twenty one" = número por extenso "tradicional").
_YEAR_WORD_PHRASES = {
    "twenty twenty": "2020", "twenty twenty one": "2021", "twenty twenty two": "2022",
    "twenty twenty three": "2023", "twenty twenty four": "2024", "twenty twenty five": "2025",
    "twenty twenty six": "2026", "twenty twenty seven": "2027",
}


def clean_year(raw):
    """Normaliza o ano-modelo para um inteiro de 4 dígitos.
    Lida com: dígitos com espaço/hífen ('20 24', '20-24'), números por
    extenso ('twenty twenty four', 'two thousand twenty one')."""
    if raw is None:
        return None
    if isinstance(raw, (int, float)):
        return int(raw)

    s = str(raw).strip().lower()

    # Dígitos puros
    if re.match(r"^\d{4}$", s):
        return int(s)

    # "20 24" ou "20-24" -> concatena os dois blocos de 2 dígitos
    m = re.match(r"^(\d{2})[\s-](\d{2})$", s)
    if m:
        return int(m.group(1) + m.group(2))

    # Frases conhecidas no estilo "twenty twenty four"
    s_norm = re.sub(r"\s+", " ", s)
    if s_norm in _YEAR_WORD_PHRASES:
        return int(_YEAR_WORD_PHRASES[s_norm])

    # Números por extenso "tradicionais": "two thousand twenty one"
    n = words_to_number(s_norm)
    if n and 1990 <= n <= 2035:
        return n

    return None


# --------------------------------------------------------------------------
# 3) Preço de venda
# --------------------------------------------------------------------------

def clean_price(raw):
    """Normaliza o preço para float (USD), tratando:
    - símbolos e sufixos: $, USD, usd, dollars
    - sufixo 'k' (milhares): '$121k' -> 121000
    - separador de milhar '.' (europeu) vs decimal ',' : '$103.750,00' -> 103750.00
    - separador de milhar ',' (americano): '$101,300.00' -> 101300.00
    - números por extenso: 'eighty two thousand USD' -> 82000.00
    """
    if raw is None:
        return None
    if isinstance(raw, (int, float)):
        return round(float(raw), 2)

    s = str(raw).strip()
    s_clean = re.sub(r"(?i)usd|dollars?|\$", "", s).strip()

    # Sufixo 'k' = milhares (checar ANTES do teste genérico de letras)
    m = re.match(r"^([\d.,]+)\s*[kK]$", s_clean)
    if m:
        num = m.group(1).replace(",", ".")
        try:
            return round(float(num) * 1000, 2)
        except ValueError:
            return None

    # Números por extenso (depois de descartar o caso do sufixo 'k')
    if re.search(r"[A-Za-z]", s_clean):
        n = words_to_number(s_clean)
        if n is not None:
            return round(float(n), 2)
        return None

    has_dot = "." in s_clean
    has_comma = "," in s_clean
    try:
        if has_dot and has_comma:
            # A ordem dos separadores define o formato:
            # vírgula antes do ponto -> EUA (milhar ',' / decimal '.')
            # ponto antes da vírgula -> Europeu (milhar '.' / decimal ',')
            if s_clean.rfind(",") < s_clean.rfind("."):
                num = s_clean.replace(",", "")
            else:
                num = s_clean.replace(".", "").replace(",", ".")
            return round(float(num), 2)
        elif has_comma and not has_dot:
            # Vírgula como separador de milhar (formato americano)
            num = s_clean.replace(",", "")
            return round(float(num), 2)
        elif has_dot and not has_comma:
            # Ambíguo: ponto pode ser milhar (3 dígitos após) ou decimal
            parts = s_clean.split(".")
            if len(parts[-1]) == 3:
                num = s_clean.replace(".", "")
            else:
                num = s_clean
            return round(float(num), 2)
        else:
            return round(float(s_clean), 2)
    except ValueError:
        return None


# --------------------------------------------------------------------------
# 4) Quilometragem do veículo
# --------------------------------------------------------------------------

_KM_TO_MILES = 0.621371


def clean_mileage(raw):
    """Normaliza a quilometragem para um inteiro em milhas, tratando:
    - unidades: 'mi', 'mi.', 'miles', 'KM'
    - separador de milhar '.' ou ',' 
    - valores por extenso: 'fifteen thousand miles' -> 15000
    - termos especiais: 'new', 'new car', 'zero', 'zero miles' -> 0
    - conversão de quilômetros para milhas quando prefixado com 'KM'
    """
    if raw is None:
        return None
    if isinstance(raw, (int, float)):
        # Valor numérico "puro": decimais pequenos são arredondados;
        # inteiros são mantidos como estão.
        return round(float(raw))

    s = str(raw).strip()
    s_low = s.lower()

    if s_low in ("new", "new car", "zero", "zero miles", "0 mi", "0 miles"):
        return 0

    # Prefixo/menção a quilômetros -> converter para milhas
    if "km" in s_low:
        digits = re.sub(r"[^\d.,]", "", s)
        digits = digits.replace(",", "")
        try:
            km = float(digits)
            return round(km * _KM_TO_MILES)
        except ValueError:
            return None

    # Remove qualquer rótulo textual (miles, mile, mi, mi., "Miles:", etc.)
    # eliminando todas as letras e dois-pontos, depois extraindo o núcleo
    # numérico remanescente — evita depender de casar cada abreviação.
    s_letters_removed = re.sub(r"(?i)[a-z:]", "", s)
    m_num = re.search(r"\d[\d.,]*", s_letters_removed)
    if m_num is None:
        # Sem dígitos: pode ser um valor totalmente por extenso
        # (remove antes as unidades "miles"/"mile"/"mi" para não
        # atrapalhar o parser de números por extenso)
        s_words_only = re.sub(r"(?i)\bmiles?\b|\bmi\b", "", s)
        if re.search(r"[A-Za-z]", s_words_only):
            return words_to_number(s_words_only)
        return None
    s_num = m_num.group(0).rstrip(".,")

    has_dot = "." in s_num
    has_comma = "," in s_num
    try:
        if has_comma:
            # vírgula = separador de milhar
            val = float(s_num.replace(",", ""))
            return round(val)
        if has_dot:
            parts = s_num.split(".")
            if len(parts[-1]) == 3:
                # ponto como separador de milhar
                val = float(s_num.replace(".", ""))
            else:
                # decimal "de verdade" -> arredonda
                val = float(s_num)
            return round(val)
        return round(float(s_num))
    except ValueError:
        return None


# --------------------------------------------------------------------------
# 5) Método de pagamento
# --------------------------------------------------------------------------

_PAYMENT_MAP = {
    "ach payment": "ACH Payment",
    "bank transfer": "Bank Transfer", "bank-transfer": "Bank Transfer", "bank_transfer": "Bank Transfer",
    "bank wire": "Wire Transfer",
    "cash": "Cash", "cash payment": "Cash",
    "credit card": "Credit Card", "creditcard": "Credit Card", "credit": "Credit Card",
    "credit card payment": "Credit Card",
    "crypto": "Crypto Payment", "crypto payment": "Crypto Payment",
    "debit card": "Debit Card",
    "financing": "Financing", "financing plan": "Financing", "finance": "Financing",
    "leasing": "Lease", "lease": "Lease", "lease plan": "Lease",
    "wire transfer": "Wire Transfer", "wiretransfer": "Wire Transfer",
    "wire": "Wire Transfer", "wire-transfer": "Wire Transfer",
}


def clean_payment_method(raw):
    if raw is None:
        return None
    key = re.sub(r"\s+", " ", str(raw).strip().lower())
    return _PAYMENT_MAP.get(key)


# --------------------------------------------------------------------------
# 6) Status de entrega
# --------------------------------------------------------------------------

_DELIVERY_MAP = {
    "cancelled": "Cancelled",
    "deliverd": "Delivered", "delivered": "Delivered",
    "in transit": "In Transit", "in-transit": "In Transit",
    "pending": "Pending",
    "awaiting delivery": "Awaiting Delivery",
    "awaiting pickup": "Awaiting Pickup",
    "awaiting review": "Awaiting Review",
    "pending approval": "Pending Approval",
    "pending review": "Pending Review",
    "shipped": "Shipped",
}


def clean_delivery_status(raw):
    if raw is None:
        return None
    key = str(raw).strip().lower()
    key = re.sub(r"[!.]+$", "", key)  # remove pontuação final ("delivered!!!", "delivered.")
    key = re.sub(r"\s+", " ", key)
    return _DELIVERY_MAP.get(key)


# --------------------------------------------------------------------------
# 7) Cidade
# --------------------------------------------------------------------------

def clean_city(raw):
    if raw is None:
        return None
    return str(raw).strip().title()


# --------------------------------------------------------------------------
# 8) Estado (UF)
# --------------------------------------------------------------------------

_STATE_NAME_TO_ABBR = {
    "alabama": "AL", "alaska": "AK", "arizona": "AZ", "arkansas": "AR", "california": "CA",
    "colorado": "CO", "connecticut": "CT", "delaware": "DE", "florida": "FL", "georgia": "GA",
    "hawaii": "HI", "idaho": "ID", "illinois": "IL", "indiana": "IN", "iowa": "IA",
    "kansas": "KS", "kentucky": "KY", "louisiana": "LA", "maine": "ME", "maryland": "MD",
    "massachusetts": "MA", "michigan": "MI", "minnesota": "MN", "mississippi": "MS",
    "missouri": "MO", "montana": "MT", "nebraska": "NE", "nevada": "NV",
    "new hampshire": "NH", "new jersey": "NJ", "new mexico": "NM", "new york": "NY",
    "north carolina": "NC", "north dakota": "ND", "ohio": "OH", "oklahoma": "OK",
    "oregon": "OR", "pennsylvania": "PA", "rhode island": "RI", "south carolina": "SC",
    "south dakota": "SD", "tennessee": "TN", "texas": "TX", "utah": "UT", "vermont": "VT",
    "virginia": "VA", "washington": "WA", "west virginia": "WV", "wisconsin": "WI",
    "wyoming": "WY",
}
_VALID_ABBRS = set(_STATE_NAME_TO_ABBR.values())


def clean_state(raw):
    if raw is None:
        return None
    s = str(raw).strip()
    if s.upper() in _VALID_ABBRS:
        return s.upper()
    return _STATE_NAME_TO_ABBR.get(s.lower())


# --------------------------------------------------------------------------
# 9) Nomes de pessoas (cliente / vendedor)
# --------------------------------------------------------------------------

def clean_person_name(raw):
    """Padroniza nomes de pessoas para Title Case e corrige hífens
    usados no lugar de espaço (ex.: 'Daniel-Jones' -> 'Daniel Jones')."""
    if raw is None:
        return None
    s = str(raw).strip()
    s = s.replace("-", " ")
    s = re.sub(r"\s+", " ", s)
    return s.title()


def resolve_salesperson_roster(cleaned_names):
    """Reconcilia vendedores cadastrados apenas com o primeiro nome
    (ex.: 'Jessica') com o nome completo correspondente já presente na
    base ('Jessica White'), quando existe exatamente UM nome completo
    compatível. Evita 'inventar' sobrenomes: se houver ambiguidade ou
    nenhuma correspondência, mantém o nome como veio."""
    fullnames_by_first = {}
    for n in cleaned_names:
        parts = n.split()
        if len(parts) >= 2:
            fullnames_by_first.setdefault(parts[0], set()).add(n)

    resolved = []
    for n in cleaned_names:
        parts = n.split()
        if len(parts) == 1:
            candidates = fullnames_by_first.get(parts[0], set())
            if len(candidates) == 1:
                resolved.append(next(iter(candidates)))
                continue
        resolved.append(n)
    return resolved


# --------------------------------------------------------------------------
# Execução principal do agente
# --------------------------------------------------------------------------

def run_agent(input_path, output_json_path):
    wb = openpyxl.load_workbook(input_path, data_only=True)
    ws = wb["Sanitized"]
    headers = [c.value for c in ws[1]]
    idx = {h: i for i, h in enumerate(headers)}

    raw_rows = []
    for r in range(2, ws.max_row + 1):
        raw_rows.append([ws.cell(row=r, column=c).value for c in range(1, len(headers) + 1)])

    cleaned = []
    for row in raw_rows:
        rec = {
            "sale_id": row[idx["sale_id"]],
            "sale_date": clean_date(row[idx["sale_date"]]),
            "customer_name": clean_person_name(row[idx["customer_name"]]),
            "porsche_model": str(row[idx["porsche_model"]]).strip(),
            "model_year": clean_year(row[idx["model_year"]]),
            "sale_price": clean_price(row[idx["sale_price"]]),
            "vehicle_mileage": clean_mileage(row[idx["vehicle_mileage"]]),
            "payment_method": clean_payment_method(row[idx["payment_method"]]),
            "city": clean_city(row[idx["city"]]),
            "state": clean_state(row[idx["state"]]),
            "salesperson": clean_person_name(row[idx["salesperson"]]),
            "delivery_status": clean_delivery_status(row[idx["delivery_status"]]),
        }
        cleaned.append(rec)

    # Reconciliação de vendedores com nome incompleto
    resolved = resolve_salesperson_roster([r["salesperson"] for r in cleaned])
    for rec, new_name in zip(cleaned, resolved):
        rec["salesperson"] = new_name

    # ---------------- Validação contra o gabarito (*Sanitized) ----------------
    checks = [
        ("sale_date", "SaleDateSanitized", lambda v: v),
        ("model_year", "ModelYearSanitized", lambda v: str(v) if v is not None else None),
        ("sale_price", "SalesPriceSanitized",
         lambda v: f"{v:.2f}" if v is not None else None),
        ("vehicle_mileage", "VehicleMileageSanitized",
         lambda v: str(v) if v is not None else None),
        ("payment_method", "PayMethodSanitized", lambda v: v),
        ("city", "CitySanitized", lambda v: v),
        ("state", "StateSanitized", lambda v: v),
        ("delivery_status", "DeliveryStatusSanitized", lambda v: v),
    ]

    report_lines = []
    total_mismatches = 0
    for field, gt_field, fmt in checks:
        mismatches = []
        for i, (row, rec) in enumerate(zip(raw_rows, cleaned)):
            gt_raw = row[idx[gt_field]]
            gt = None if gt_raw is None else str(gt_raw).strip()
            got = fmt(rec[field])
            if gt_field == "SalesPriceSanitized" and got is not None:
                got_f = float(got)
                gt_f = float(gt) if gt else None
                ok = gt_f is not None and abs(got_f - gt_f) < 0.01
            else:
                ok = (got == gt) or (got is None and gt in (None, "", "INVALID") and field == "sale_date")
            if not ok:
                mismatches.append((row[idx["sale_id"]], gt, got))
        acc = 100 * (len(raw_rows) - len(mismatches)) / len(raw_rows)
        report_lines.append(f"{field:18s}: {acc:5.1f}% de acerto ({len(mismatches)} divergências)")
        total_mismatches += len(mismatches)
        if mismatches:
            for sid, gt, got in mismatches[:5]:
                report_lines.append(f"    sale_id={sid}: esperado={gt!r} | agente={got!r}")

    print("=" * 70)
    print("RELATÓRIO DE VALIDAÇÃO DO AGENTE (comparação com colunas *Sanitized*)")
    print("=" * 70)
    for line in report_lines:
        print(line)
    print("-" * 70)
    print(f"Total de divergências: {total_mismatches} em {len(raw_rows) * len(checks)} verificações")

    # Contagem de datas inválidas (fluxo esperado, não é erro do agente)
    n_invalid_dates = sum(1 for r in cleaned if r["sale_date"] == "INVALID")
    print(f"Datas marcadas como INVALID (irrecuperáveis): {n_invalid_dates}")

    with open(output_json_path, "w", encoding="utf-8") as f:
        json.dump(cleaned, f, ensure_ascii=False, indent=2)
    print(f"\nBase tratada salva em: {output_json_path}")

    return cleaned, report_lines


if __name__ == "__main__":
    run_agent(
        "/home/claude/porsche/base_porsche.xlsx",
        "/home/claude/porsche/dados_tratados.json",
    )
